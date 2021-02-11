import sqlite3
from openpyxl import load_workbook

def rowcount():
    count = 0
    for rowv in ws.values:
        if rowv[1] != None:
            count += 1
            continue
        else:
            return count

# create db file
connection = sqlite3.connect('test95.db')
cursor = connection.cursor()
cursor.execute('''CREATE TABLE IF NOT EXISTS "Players" (
    "id"	INTEGER,
    "name"	TEXT,
    "dob"   TEXT,
    "gender"    TEXT,
    "counter"   INTEGER PRIMARY KEY AUTOINCREMENT
);''')
cursor.execute('''CREATE TABLE IF NOT EXISTS "Seasons" (
	"id"	INTEGER,
	"season"	TEXT,
    PRIMARY KEY("id")
);''')
cursor.execute('''CREATE TABLE IF NOT EXISTS "Stats" (
	"player_id"	INTEGER,
    "name"      TEXT,
    "age_time_of_filming"	INTEGER,
    "hometown"	TEXT,
    "season_id"	INTEGER,
    "finish"	TEXT,
    "days_lasted"	INTEGER,
    "votes_against"	INTEGER,
    "tribal_wins"	INTEGER,
    "individual_wins"	INTEGER,
    "total_wins"	INTEGER,
    "jury_voter"	INTEGER,
    "final_contestant"	INTEGER,
    "sole_survivor"	INTEGER,
	"final_vote_id"	INTEGER,
    "counter"   INTEGER PRIMARY KEY AUTOINCREMENT
);''')

# load .xlsx file
wb = load_workbook(filename = 'Survivor_Contestants.xlsx')

# set to 3rd sheet (season)
wb.active = 2
ws = wb.active

# find out how many rows are in .xlsx file (r)
r = rowcount()

for row in ws.iter_rows(min_row=2, min_col=1, max_col=2, max_row=r, values_only=True):
    cursor.execute("INSERT INTO Seasons VALUES (?, ?)", row)
    connection.commit()

# set to 2nd sheet (player)
wb.active = 1
ws = wb.active

# find out how many rows are in .xlsx file (r)
r = rowcount()

for row in ws.iter_rows(min_row=2, min_col=1, max_col=4, max_row=r, values_only=True):
    cursor.execute("INSERT INTO Players (id, name, dob, gender) VALUES (?, ?, ?, ?)", row)
    connection.commit()

player_list_start = cursor.execute("SELECT name FROM Players")
saved_list = player_list_start.fetchall()

# number id should start on
id = 1

# set counter at 1
c = 1

# list of players to keep track if they are already in the table
names = []

for item in saved_list:
    if item[0] not in names:
        cursor.execute("UPDATE Players SET id = ? WHERE name = ? AND counter = ?", (id, item[0], c))
        connection.commit()
        names.append(item[0])
        id += 1
    else:
        cursor.execute("DELETE FROM Players WHERE name = ? AND counter = ?", (str(item[0]), c))
        connection.commit()
    c += 1

# set to 1st sheet (stats)
wb.active = 0
ws = wb.active

# find out how many rows are in .xlsx file (r)
r = rowcount()

for row in ws.iter_rows(min_row=3, min_col=2, max_col=15, max_row=r, values_only=True):
    cursor.execute("INSERT INTO Stats (name, age_time_of_filming, hometown, season_id, finish, days_lasted, votes_against, tribal_wins, individual_wins, total_wins, final_contestant, sole_survivor, jury_voter, final_vote_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", row)
    connection.commit()

## code dealing with players who competed under 2 different names
cursor.execute("ALTER TABLE Players ADD alt_name TEXT")
connection.commit()

# add alternate names to orginal names in Players
cursor.execute("UPDATE Players SET alt_name = 'Amber Mariano' WHERE name = 'Amber Brkich'")
connection.commit()
cursor.execute("UPDATE Players SET alt_name = 'Kim Spradlin-Wolfe' WHERE name = 'Kim Spradlin'")
connection.commit()
cursor.execute("UPDATE Players SET alt_name = 'Candice Cody' WHERE name = 'Candice Woodcock'")
connection.commit()

# find the second names in Players, delete them
cursor.execute("DELETE FROM Players WHERE name = 'Amber Mariano'")
connection.commit()
cursor.execute("DELETE FROM Players WHERE name = 'Kim Spradlin-Wolfe'")
connection.commit()
cursor.execute("DELETE FROM Players WHERE name = 'Candice Cody'")
connection.commit()

# set counter at 1
c = 1

list_length = len(cursor.execute("SELECT name FROM Stats").fetchall())

while c <= list_length:
    stat_name = cursor.execute("SELECT name FROM Stats WHERE counter = ?", (str(c),)).fetchone()[0]
    pull_id = cursor.execute("SELECT id FROM Players WHERE name = ? OR alt_name = ?", (stat_name, stat_name)).fetchone()
    str_id = str(pull_id).strip("(),")
    cursor.execute("UPDATE Stats SET player_id = ? WHERE name = ?", (str_id, str(stat_name)))
    connection.commit()
    c += 1

# TODO - delete the name column in Stats
# cursor.execute("CREATE TABLE Stats_temp AS SELECT player_id, age_time_of_filming, hometown, season_id, finish, days_lasted, votes_against, tribal_wins, individual_wins, total_wins, jury_voter, final_contestant, sole_survivor, final_vote_id FROM Stats")
# connection.commit()
# cursor.execute("DROP Stats")
# connection.commit()
# cursor.execute("ALTER TABLE Stats_temp RENAME TO Stats")
# connection.commit()
#
# cursor.execute("CREATE TABLE Players_temp AS SELECT id, name, dob, gender FROM Players")
# connection.commit()
# cursor.execute("DROP Players")
# connection.commit()
# cursor.execute("ALTER TABLE Players_temp RENAME TO Players")
# connection.commit()
